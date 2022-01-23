from rest_framework.response import Response 
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from django.http import HttpResponse
from base.models import Contact
from .serializers import ContactSerializer, RegistrationSerializer
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from django.core.mail import send_mail


class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        # Add custom claims
        token['username'] = user.username

        return token

class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

@api_view(['GET'])
def getRoutes(request):
    routes = [
        '/api/token',
        '/api/toke/refresh',
    ]
    return Response(routes)

@api_view(['POST'])
def registration_view(request):
    serializer = RegistrationSerializer(data=request.data)
    data = {}
    if serializer.is_valid():
        user = serializer.save()
        data['response'] = "successfully registered a new user"
        data['email'] = user.email
        data['username'] = user.username
    else:
        data = serializer.errors
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_file_view(request):
    folder='files/'
    myfile = request.FILES['File']
    fs = FileSystemStorage(location=folder)
    filename = fs.save(myfile.name, myfile)
    file_url = fs.url(filename)
    wb = load_workbook(filename = folder + filename)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        contact = Contact()
        contact.user = request.user
        contact.first_name = ws[form_cell_name(row, 1)].value
        contact.last_name = ws[form_cell_name(row, 2)].value
        contact.email = ws[form_cell_name(row, 3)].value
        contact.address_line_1 = ws[form_cell_name(row, 4)].value
        contact.address_line_2 = ws[form_cell_name(row, 5)].value
        contact.city = ws[form_cell_name(row, 6)].value
        contact.state = ws[form_cell_name(row, 7)].value
        if ws[form_cell_name(row, 8)].value:
            contact.zipcode = int(ws[form_cell_name(row, 8)].value)
        contact.picture_url = ws[form_cell_name(row, 9)].value
        contact.save()
    return Response({
        'file_url': file_url
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_file_view(request):
    folder='files/'
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    for contact in request.user.contact_set.all():
        
        contact_data = []
        contact_data.append(contact.first_name)
        contact_data.append(contact.last_name)        
        contact_data.append(contact.email)
        contact_data.append(contact.address_line_1)
        contact_data.append(contact.address_line_2)
        contact_data.append(contact.city)
        contact_data.append(contact.state)
        contact_data.append(contact.zipcode)
        contact_data.append(contact.picture_url)
        ws.append(contact_data)
    wb.save(filename = folder + "output.xlsx")
    file = open(folder + "output.xlsx", 'rb')

    response = HttpResponse(file, content_type='"application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=myfile.xlsx'

    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def send_postcard_view(request):
    for contact in request.user.contact_set.all():
        if contact.email:
            send_mail('Postcard title', 'Postcard message', contact.email, [contact.email])

    return Response("mails sended")

def form_cell_name(row, col):
    return get_column_letter(col) + str(row)

@api_view(['GET', 'POST', 'PUT'])
@permission_classes([IsAuthenticated])
def getContacts(request):
    if request.method == 'GET':
        user = request.user
        contacts = user.contact_set.all()
        serializer = ContactSerializer(contacts, many=True)
        return Response(serializer.data)
    elif request.method == 'POST':
        serializer = ContactSerializer(data=request.data)
        if serializer.is_valid():
            contact = serializer.save(request.user)
            return Response("success")
        else:
            return Response(serializer.errors)
    elif request.method == 'PUT':
        serializer = ContactSerializer(data=request.data)
        if serializer.is_valid():
            contact = serializer.update()
            return Response("updated")
        else:
            return Response(serializer.errors)

@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def getContactsId(request, id):
    if request.method == 'GET':
        user = request.user
        contact = Contact.objects.filter(pk=id).first()
        serializer = ContactSerializer(contact)
        return Response(serializer.data)
    elif request.method == 'DELETE':
        contact = Contact.objects.filter(pk=id)
        contact.delete()
        return Response("success")